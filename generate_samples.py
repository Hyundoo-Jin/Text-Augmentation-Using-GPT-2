import argparse, json, os, sys, time, tqdm, pickle
import numpy as np
import tensorflow as tf
from tensorflow.core.protobuf import rewriter_config_pb2
import sentencepiece as spm
from load_dataset import load_dataset, Sampler

import encoder, sample_cls
import transfer_clstoken as model
from accumulate import AccumulatingOptimizer
import memory_saving_gradients
from openpyxl import Workbook, load_workbook



CHECKPOINT_DIR = 'checkpoint'
SAMPLE_DIR = 'samples'

parser = argparse.ArgumentParser(
    description = '과제 데이터 fine tuning argument',
    formatter_class = argparse.ArgumentDefaultsHelpFormatter
)

parser.add_argument('--batch_size', metavar = 'SIZE', type = int, default = 1, help = 'Batch size')
parser.add_argument('--learning_rate', metavar = 'LR', type = float, default = 0.00002, help = 'Learning rate for Adam')
parser.add_argument('--accumulate_gradients', metavar = 'N', type = int, default = 1, help = 'Accumulate gradients across N minibatches')
parser.add_argument('--memory_saving_gradients', default = False, action = 'store_true', help = 'Use gradient checkpointing to reduce vram usage.')
parser.add_argument('--only_train_transformer_layers', default = False, action = 'store_true', help = 'Restrict training to the transformer blocks.')
parser.add_argument('--optimizer', type = str, default = 'adam', help = 'Optimizer <adam | sgd>')
parser.add_argument('--noise', type = float, default = 0.0, help = 'Add noise to input training data to regularize against typos')
parser.add_argument('--encoder_path', type = str, metavar = 'PATH', default = '../long_train.model', help = 'Sentencepiece model path')

parser.add_argument('--top_k', type = int, default = 40, help = 'k for top-k sampling')
parser.add_argument('--top_p', type = float, default = 0.0, help = 'p for top-p sampling')

parser.add_argument('--restore_from', type = str, default = 'lastest', help = 'Either "lastest", "fresh", or a path to a checkpoint file')
parser.add_argument('--run_name', type = str, default = 'no_classification', help = 'Run id. Name of subdirectory in checkpoint/ and samples/')
parser.add_argument('--n_ctx', metavar = 'N', type = int, default = 32, help='number of input tokens')
parser.add_argument('--temperature', type = float, default = 1.0)
parser.add_argument('--n_samples', type = int, default = 300)
parser.add_argument('--result_file', type = str, metavar = 'PATH', default = 'generated.xlsx')

def maketree(path) :
    try :
        os.makedirs(path)
    except :
        pass

def randomize(context, hparams, p) :
    if p > 0 :
        mask = tf.random.uniform(shape = tf.shape(context)) < p
        noise = tf.random.uniform(shape = tf.shape(context), minval = 0, maxval = hparams.n_vocab, dtype = tf.int32)
        return tf.where(mask, noise, context)
    else :
        return context

def main() :
    args = parser.parse_args()
    config = tf.ConfigProto()
    config.gpu_options.allow_growth = True
    config.graph_options.rewrite_options.layout_optimizer = rewriter_config_pb2.RewriterConfig.OFF
    sp = spm.SentencePieceProcessor()
    sp.load(args.encoder_path)
    with tf.Session(config = config) as sess :
        hparams = model.default_hparams()
        context = tf.placeholder(tf.int32, [1, None])
        labels = tf.placeholder(tf.int32, [1, None])
        with open('hparams.json') as f :
            hparams.override_from_dict(json.load(f))
        hparams.n_vocab = hparams.n_vocab + 9   # cls 토큰 수 추가
        tf_sample = sample_cls.sample_sequence(
            hparams = hparams,
            length = 64,
            context = context,
            labels = labels,
            batch_size = 1,
            temperature = 1.0,
            top_k = args.top_k,
            top_p = args.top_p
        )
        with open('../cls_generate.pkl', 'rb') as f :
            chunks, q_ids = pickle.load(f)
        data_sampler = Sampler(chunks, transfer = True, labels=True)
        data_sampler.add_ids(np.array(q_ids))

        saver = tf.train.Saver()
        ckpt = tf.train.latest_checkpoint(
            os.path.join(CHECKPOINT_DIR, args.run_name))
        saver.restore(sess, ckpt)
        result = load_workbook('template.xlsx')
        counter = [0] * 55
        with open('qname_dict.json', 'r') as f :
            q_dict = json.load(f)
        while True :
            try :
                for q_id in range(max(q_ids) + 1) :
                    sheet_name = q_dict[str(q_id)]['q_name']
                    n_answers = q_dict[str(q_id)]['n_answers']
                    if str(q_id) not in ['11', '12', '6', '7', '19', '2', '3', '8', '0', '13', '14', '5', '26', '27', '29', '10', '15', '16'] : continue
                    for _ in range(int(args.n_samples / args.batch_size) + 1) :
                        batch = data_sampler.sample_simple(q_id, args.batch_size, generate = True, after = True)
                        for context_token, context_label in batch :
                            ############### 오답만 생성하기 위한 처리
#                             context_label = list()
                            ############### 오답만 생성하기 위한 처리
                            out = sess.run(tf_sample, 
                                feed_dict={context : [context_token],
                                           labels : [context_label]})
                            out = out.tolist()[0]
                            o = [idx for idx in out if idx not in [10000, 10001, 10002, 10003, 10004, 10005, 10006, 10007, 10008]]
                            two = o.index(2)
                            q = o[:two]
                            a = o[two+1:]
                            question = ''
                            answer = ''
                            question += sp.DecodeIds(q)
                            answer += sp.DecodeIds(a).split('<|')[0]
                            counter[q_id] += 1
                            result[sheet_name].cell(row = counter[q_id] + 1, column = 1).value = answer
                            for lbl in range(10000, 10000 + n_answers) :
                                if lbl in context_label :
                                    result[sheet_name].cell(row = counter[q_id] + 1, column = lbl-9998).value = 1
                                else :
                                    result[sheet_name].cell(row = counter[q_id] + 1, column = lbl-9998).value = 0
                            print('Question : ', question, 'Answer : ', answer)
                result.save(filename=args.result_file) 
                print('샘플링 종료')
                break
            except KeyboardInterrupt :
                result.save(filename=args.result_file)
                print('샘플링 종료')
                break

if __name__ == '__main__' :
    main()
